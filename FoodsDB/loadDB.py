# Import database module.
import firebase_admin
from firebase_admin import db, credentials
import openpyxl
from openpyxl import load_workbook


tagMeaning={
        "PROCNT": "Protein",
        "FAT": "Fat",
        "CHOCDF": "Carbohydrate",
        "ENERC_KCAL": "Calorie",
        "LACS": "Lactose",
        "ALC": "Alcohol",
        "CAFFN": "Caffeine",
        "FIBTG": "Fiber",
        "CA": "Calcium",
        "FE": "Iron",
        "NA": "Sodium",
        "VITA_RAE": "Vitamin A",
        "TOCPHA": "Vitamin E",
        "VITD" : "Vitamin D",
        "VITC" : "Vitamin C",
        "VITB6A" : "Vitamin B-6",
        "VITB12" : "Vitamin B-12",
        "FOLAC" : "Folic acid",
        "CHOLE" : "Cholesterol"
    }

cred = credentials.Certificate('E:\Technion\semester 8\scripts\\team5fooddb.json')

# Get a database reference to our blog.
firebase_admin = firebase_admin.initialize_app(cred, {'databaseURL': 'https://team5fooddb.firebaseio.com'})
ref = db.reference('')

foods_ref = ref.child('food')

def isValid(cfood):
    if "$" in cfood or "#" in cfood or "[" in cfood or "]" in cfood or "/" in cfood or "." in cfood:
        return False
    return True

def remove_spaces(str):
    if len(str)== 0:
        return str
    start = 0
    end = len(str)
    if str[0] is ' ':
        start = 1
    if str[end-1] is ' ':
        end = end - 1
    return str[start:end]


def tag2string(tag):
    return tagMeaning[tag]


def edit_name(food):
    splited = food.split(",")
    to_ret = ''
    genFood=splited[0]
    if "with" in splited[len(splited) - 1]:
        to_ret = splited[len(splited) - 1]
        del splited[(len(splited) - 1)]
    for s in splited:
        to_ret = remove_spaces(s) + " " + to_ret
    return (genFood+"-"+remove_spaces(to_ret)).upper()


wb = load_workbook(filename='foods.xlsx')
ws = wb.active

dict = {}
cfood = ""

i = 2
while ws["A" + str(i)].value is not None:
    if edit_name(ws["A" + str(i)].value) != cfood:
        cfood = edit_name(ws["A" + str(i)].value)
        if not isValid(cfood):
            i=i+1
            continue
        dict[cfood] = {}
    if isValid(cfood):
        dict[cfood][tag2string(ws['B' + str(i)].value)] = ws['C' + str(i)].value
    i = i + 1

#print(dict)
foods_ref.set(dict)
#foods_ref.set({})
